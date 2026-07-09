"""Proyeksi Kebutuhan Nasional SDM Jabatan Fungsional Arsiparis (JFA)."""
from __future__ import annotations

import json, math, re
from datetime import date
from pathlib import Path
import pandas as pd

try:
    from scipy import stats as scipy_stats
except Exception:
    scipy_stats = None

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "proyeksi_kebutuhan_sdm_arsiparis"
OUT_DIR.mkdir(parents=True, exist_ok=True)

JENJANG = [
    ("terampil", "Terampil"), ("mahir", "Mahir"), ("penyelia", "Penyelia"),
    ("ahli_pertama", "Ahli Pertama"), ("ahli_muda", "Ahli Muda"),
    ("ahli_madya", "Ahli Madya"), ("ahli_utama", "Ahli Utama"),
]

OBSERVATIONS = pd.DataFrame([
    {"periode":"2025-06", "tanggal":date(2025,6,1), "bulan_ke":0, "total":23338},
    {"periode":"2026-05", "tanggal":date(2026,5,1), "bulan_ke":11, "total":27875},
    {"periode":"2026-06", "tanggal":date(2026,6,1), "bulan_ke":12, "total":27820},
    {"periode":"2026-07", "tanggal":date(2026,7,1), "bulan_ke":13, "total":27837},
])

OFFICIAL_DASHBOARD = pd.DataFrame([
    {"key":"terampil", "jenjang":"Terampil", "kebutuhan":35679, "tersedia":7005, "kekurangan":28674},
    {"key":"mahir", "jenjang":"Mahir", "kebutuhan":24420, "tersedia":906, "kekurangan":23514},
    {"key":"penyelia", "jenjang":"Penyelia", "kebutuhan":17488, "tersedia":594, "kekurangan":16894},
    {"key":"ahli_pertama", "jenjang":"Ahli Pertama", "kebutuhan":29558, "tersedia":13546, "kekurangan":16012},
    {"key":"ahli_muda", "jenjang":"Ahli Muda", "kebutuhan":18837, "tersedia":4650, "kekurangan":14187},
    {"key":"ahli_madya", "jenjang":"Ahli Madya", "kebutuhan":4602, "tersedia":1094, "kekurangan":3508},
    {"key":"ahli_utama", "jenjang":"Ahli Utama", "kebutuhan":132, "tersedia":25, "kekurangan":107},
])
IDEAL_TOTAL = int(OFFICIAL_DASHBOARD.kebutuhan.sum())
DASHBOARD_AVAILABLE_TOTAL = int(OFFICIAL_DASHBOARD.tersedia.sum())


def normalize_instansi(value: str) -> str:
    text = str(value or "").upper()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[/.,;:\-]+", " ", text)
    text = re.sub(r"\bPEMERINTAH\s+KAB\b", "KABUPATEN", text)
    text = re.sub(r"\bPEMERINTAH\s+KABUPATEN\b", "KABUPATEN", text)
    text = re.sub(r"\bPEMERINTAH\s+KOTA\b", "KOTA", text)
    text = re.sub(r"\bPEMERINTAH\s+PROVINSI\b", "PROVINSI", text)
    text = re.sub(r"\s+", " ", text).strip()
    aliases = {
        "KEJAKSAAN AGUNG":"KEJAKSAAN REPUBLIK INDONESIA",
        "KEJAKSAAN RI":"KEJAKSAAN REPUBLIK INDONESIA",
        "PROVINSI DKI JAKARTA":"PROVINSI DAERAH KHUSUS IBUKOTA JAKARTA",
    }
    return aliases.get(text, text)


def read_rekap_excel(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, header=None)
    header_row = None
    for i in range(min(30, len(raw))):
        joined = " ".join(str(x).strip().lower() for x in raw.iloc[i].tolist())
        if "nama" in joined and "instansi" in joined and "total" in joined:
            header_row = i; break
    if header_row is None:
        raise ValueError(f"Header rekap tidak ditemukan: {path}")
    df = pd.read_excel(path, skiprows=header_row)
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    if "nama_instansi" not in df.columns:
        candidates = [c for c in df.columns if "instansi" in c]
        if candidates: df = df.rename(columns={candidates[0]:"nama_instansi"})
    df = df[df.nama_instansi.notna()].copy()
    df = df[~df.nama_instansi.astype(str).str.upper().str.contains("TOTAL KESELURUHAN", na=False)]
    for col, _ in JENJANG:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    if "total" in df.columns: df["total"] = pd.to_numeric(df.total, errors="coerce").fillna(0).astype(int)
    df["nama_normal"] = df.nama_instansi.map(normalize_instansi)
    return df


def extract_balanced_js_object(text: str, marker: str = "const RAW") -> dict:
    pos = text.find(marker)
    if pos < 0: raise ValueError(f"Marker {marker!r} tidak ditemukan")
    start = text.find("{", pos)
    depth = 0; in_str = False; quote = ""; escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if escape: escape = False
            elif ch == "\\": escape = True
            elif ch == quote: in_str = False
        else:
            if ch in ("'", '"'): in_str = True; quote = ch
            elif ch == "{": depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0: return json.loads(text[start:i+1])
    raise ValueError("Object JS tidak tertutup sempurna")


def parse_bkn_raw_html(path: Path) -> pd.DataFrame:
    raw = extract_balanced_js_object(path.read_text(encoding="utf-8", errors="ignore"))
    rows = []
    for ins_i, jab_i, tmt_i, tingkat_i, pend_i, asn_i, jumlah in raw["rows"]:
        rows.append({
            "instansi": raw["ins_r"][str(ins_i)], "jabatan": raw["jab_r"][str(jab_i)],
            "tmt_jenjang": raw["tmt_r"][str(tmt_i)], "tingkat_pendidikan": raw["tingkat_r"][str(tingkat_i)],
            "jurusan": raw["pend_r"][str(pend_i)], "asn": raw["asn_r"][str(asn_i)], "jumlah": int(jumlah)
        })
    return pd.DataFrame(rows)


def tcrit_975(df: int) -> float:
    if scipy_stats is not None: return float(scipy_stats.t.ppf(0.975, df))
    if df == 2: return 4.302652729911275
    raise RuntimeError("scipy dibutuhkan untuk df selain 2")


def two_sided_pvalue_t(t_abs: float, df: int) -> float:
    if scipy_stats is not None: return float(2 * scipy_stats.t.sf(t_abs, df))
    if df == 2:
        cdf = 0.5 + t_abs / (2 * math.sqrt(t_abs*t_abs + 2))
        return float(2 * (1 - cdf))
    raise RuntimeError("scipy dibutuhkan untuk df selain 2")


def fit_ols(df: pd.DataFrame) -> tuple[dict, pd.DataFrame]:
    x = df.bulan_ke.astype(float); y = df.total.astype(float); n = len(df)
    xbar = x.mean(); ybar = y.mean()
    sxx = float(((x-xbar)**2).sum()); sxy = float(((x-xbar)*(y-ybar)).sum())
    slope = sxy / sxx; intercept = ybar - slope*xbar; yhat = intercept + slope*x
    resid = y - yhat; sse = float((resid**2).sum()); sst = float(((y-ybar)**2).sum())
    df_resid = n - 2; mse = sse / df_resid; rse = math.sqrt(mse); se_slope = math.sqrt(mse/sxx)
    tcrit = tcrit_975(df_resid); pval = two_sided_pvalue_t(abs(slope/se_slope), df_resid)
    result = {"intercept":float(intercept), "slope":float(slope), "r_squared":float(1-sse/sst),
              "p_value_slope":pval, "std_err_slope":float(se_slope),
              "slope_ci_low":float(slope-tcrit*se_slope), "slope_ci_high":float(slope+tcrit*se_slope),
              "residual_std_error":float(rse), "t_critical_95":tcrit, "df":df_resid,
              "n_observasi":n}
    fitted = df.copy(); fitted["fitted_total"] = yhat; fitted["residual"] = resid
    return result, fitted


def projection_ci(x0: float, result: dict, observed: pd.DataFrame) -> tuple[float, float, float]:
    x = observed.bulan_ke.astype(float); n = len(x); xbar = x.mean(); sxx = float(((x-xbar)**2).sum())
    yhat = result["intercept"] + result["slope"]*x0
    se_mean = result["residual_std_error"] * math.sqrt((1/n) + ((x0-xbar)**2/sxx))
    margin = result["t_critical_95"] * se_mean
    return float(yhat), float(yhat-margin), float(yhat+margin)


def allocate_by_largest_remainder(total: float, proportions: pd.Series) -> dict:
    target = int(round(total)); raw = proportions * target
    base = raw.apply(math.floor).astype(int); need = target - int(base.sum())
    for key in (raw-base).sort_values(ascending=False).index[:need]: base.loc[key] += 1
    return base.to_dict()


def build_projection(result: dict) -> pd.DataFrame:
    props = OFFICIAL_DASHBOARD.set_index("key").tersedia / DASHBOARD_AVAILABLE_TOTAL
    rows = []
    for year in range(2026, 2032):
        bulan_ke = (year - 2025) * 12 + 1  # Juli setiap tahun.
        est, lo, hi = projection_ci(bulan_ke, result, OBSERVATIONS)
        alloc = allocate_by_largest_remainder(est, props)
        row = {"tahun":year, "bulan_ke":bulan_ke, "total_estimasi":round(est), "total_ci_bawah":round(lo), "total_ci_atas":round(hi)}
        row.update({label:alloc[key] for key, label in JENJANG})
        rows.append(row)
    return pd.DataFrame(rows)


def build_gap_2031(projection: pd.DataFrame) -> pd.DataFrame:
    r = projection[projection.tahun == 2031].iloc[0]
    scenarios = [("Batas bawah 95% CI", int(r.total_ci_bawah)), ("Titik estimasi", int(r.total_estimasi)), ("Batas atas 95% CI", int(r.total_ci_atas))]
    return pd.DataFrame([{"skenario_2031":n, "proyeksi_tersedia":v, "kebutuhan_ideal":IDEAL_TOTAL,
                          "gap":IDEAL_TOTAL-v, "persen_belum_terpenuhi":(IDEAL_TOTAL-v)/IDEAL_TOTAL} for n,v in scenarios])


def fmt_int(v): return f"{int(round(v)):,}".replace(",", ".")
def fmt_float(v, d=3): return f"{v:,.{d}f}".replace(",", "X").replace(".", ",").replace("X", ".")



def df_to_markdown(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    rows = []
    rows.append("| " + " | ".join(str(c) for c in cols) + " |")
    rows.append("| " + " | ".join("---" for _ in cols) + " |")
    for _, row in df.iterrows():
        vals = []
        for c in cols:
            v = row[c]
            if isinstance(v, float):
                vals.append(fmt_float(v, 3) if abs(v) < 10 else fmt_float(v, 1))
            else:
                vals.append(str(v))
        rows.append("| " + " | ".join(vals) + " |")
    return "\n".join(rows)

def write_markdown(result: dict, projection: pd.DataFrame, gap: pd.DataFrame) -> None:
    est = int(gap.loc[gap.skenario_2031 == "Titik estimasi", "proyeksi_tersedia"].iloc[0])
    g = int(gap.loc[gap.skenario_2031 == "Titik estimasi", "gap"].iloc[0])
    pct = float(gap.loc[gap.skenario_2031 == "Titik estimasi", "persen_belum_terpenuhi"].iloc[0])
    lo = int(gap.loc[gap.skenario_2031 == "Batas bawah 95% CI", "proyeksi_tersedia"].iloc[0])
    hi = int(gap.loc[gap.skenario_2031 == "Batas atas 95% CI", "proyeksi_tersedia"].iloc[0])
    md = f"""# Proyeksi Kebutuhan Nasional SDM Jabatan Fungsional Arsiparis

## Statistik Regresi OLS

- Slope: {fmt_float(result['slope'], 3)} orang/bulan
- Intercept: {fmt_float(result['intercept'], 3)}
- R-squared: {fmt_float(result['r_squared'], 4)}
- p-value slope: {fmt_float(result['p_value_slope'], 4)}
- Standard error slope: {fmt_float(result['std_err_slope'], 3)}
- 95% CI slope: {fmt_float(result['slope_ci_low'], 3)} s.d. {fmt_float(result['slope_ci_high'], 3)} orang/bulan

## Ringkasan Naratif

Berdasarkan empat titik data nasional, yaitu Juni 2025, Mei 2026, Juni 2026, dan Juli 2026, model regresi linear sederhana menghasilkan kecenderungan kenaikan sebesar {fmt_float(result['slope'], 1)} arsiparis per bulan. Namun demikian, interpretasi tren perlu dilakukan secara hati-hati karena jumlah observasi sangat terbatas dan tiga titik terakhir berada dalam rentang waktu yang berdekatan. Dengan R-squared sebesar {fmt_float(result['r_squared'], 3)}, model ini terutama berguna sebagai baseline indikatif, bukan sebagai angka prediksi presisi.

Jika kecenderungan linear tersebut digunakan untuk proyeksi sampai Juli 2031, jumlah arsiparis nasional diperkirakan mencapai {fmt_int(est)} orang, dengan rentang confidence interval 95% sekitar {fmt_int(lo)} sampai {fmt_int(hi)} orang. Dibandingkan dengan kebutuhan ideal berdasarkan rekomendasi dashboard sebesar {fmt_int(IDEAL_TOTAL)} orang, masih terdapat gap sekitar {fmt_int(g)} orang atau {fmt_float(pct*100, 1)}% dari kebutuhan ideal pada titik estimasi 2031.

Dengan demikian, pertumbuhan alamiah berdasarkan tren historis singkat belum memadai untuk menutup kekurangan nasional JFA dalam lima tahun. Selain itu, coverage rekomendasi baru sekitar 46% nasional, sehingga angka kebutuhan ideal dan kekurangan yang digunakan dalam pembandingan ini masih berpotensi underestimate terhadap kebutuhan riil, terutama pada pemerintah kabupaten/kota yang cakupan rekomendasinya relatif lebih rendah.

## Keterbatasan Metodologis

- Model hanya memakai empat titik waktu sehingga parameter regresi memiliki ketidakpastian tinggi.
- Asumsi linearitas belum tentu menangkap pola non-linear akibat siklus formasi CPNS/PPPK, anggaran, dan kebijakan pembinaan JFA.
- Pembagian proyeksi per jenjang memakai proporsi komposisi dashboard Juli 2026 sebagai simplifying assumption.
- Data BKN 2025 tidak memuat usia/tanggal lahir, sehingga proyeksi pensiun berbasis usia tidak dihitung.
- Perubahan nomenklatur/restrukturisasi kementerian dapat memengaruhi perbandingan per instansi dan tidak boleh langsung dibaca sebagai rekrutmen bersih.

## Tabel Proyeksi Tahunan

{df_to_markdown(projection)}

## Gap 2031 terhadap Kebutuhan Ideal

{df_to_markdown(gap.assign(persen_belum_terpenuhi=lambda d: (d.persen_belum_terpenuhi*100).round(1)))}
"""
    (OUT_DIR / "ringkasan_naratif_proyeksi.md").write_text(md, encoding="utf-8")


def write_svg_chart(projection: pd.DataFrame) -> None:
    width, height = 920, 520; pl, pr, pt, pb = 80, 40, 50, 75
    xmin, xmax = 0, int(projection.bulan_ke.max())
    ymin = min(int(OBSERVATIONS.total.min()), int(projection.total_ci_bawah.min())) - 1500
    ymax = max(int(OBSERVATIONS.total.max()), int(projection.total_ci_atas.max())) + 1500
    sx = lambda x: pl + (x-xmin)/(xmax-xmin)*(width-pl-pr)
    sy = lambda y: height-pb - (y-ymin)/(ymax-ymin)*(height-pt-pb)
    upper = " ".join(f"{sx(r.bulan_ke):.1f},{sy(r.total_ci_atas):.1f}" for r in projection.itertuples())
    lower = " ".join(f"{sx(r.bulan_ke):.1f},{sy(r.total_ci_bawah):.1f}" for r in reversed(list(projection.itertuples())))
    proj_line = " ".join(f"{sx(r.bulan_ke):.1f},{sy(r.total_estimasi):.1f}" for r in projection.itertuples())
    actual_line = " ".join(f"{sx(r.bulan_ke):.1f},{sy(r.total):.1f}" for r in OBSERVATIONS.itertuples())
    grid = "".join(f'<line x1="{pl}" y1="{sy(y):.1f}" x2="{width-pr}" y2="{sy(y):.1f}" stroke="#e7edf3"/><text x="{pl-12}" y="{sy(y)+4:.1f}" text-anchor="end" font-size="12" fill="#445">{fmt_int(y)}</text>' for y in range((ymin//5000+1)*5000, ymax+1, 5000))
    labels = "".join(f'<text x="{sx(r.bulan_ke):.1f}" y="{height-38}" text-anchor="middle" font-size="12" fill="#445">{r.tahun}</text>' for r in projection.itertuples())
    points = "".join(f'<circle cx="{sx(r.bulan_ke):.1f}" cy="{sy(r.total):.1f}" r="5" fill="#0b5cab"/>' for r in OBSERVATIONS.itertuples())
    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
<rect width="100%" height="100%" fill="#fff"/><text x="{pl}" y="30" font-family="Arial" font-size="20" font-weight="700" fill="#18212f">Proyeksi Nasional SDM JFA 2026-2031</text>
<text x="{pl}" y="50" font-family="Arial" font-size="12" fill="#526070">OLS total nasional; area biru muda menunjukkan 95% confidence interval mean projection.</text><g font-family="Arial">{grid}</g>
<line x1="{pl}" y1="{height-pb}" x2="{width-pr}" y2="{height-pb}" stroke="#9aa8b6"/><line x1="{pl}" y1="{pt}" x2="{pl}" y2="{height-pb}" stroke="#9aa8b6"/>
<polygon points="{upper} {lower}" fill="#8ec5ff" opacity="0.28"/><polyline points="{proj_line}" fill="none" stroke="#0f766e" stroke-width="3"/><polyline points="{actual_line}" fill="none" stroke="#0b5cab" stroke-width="3"/>{points}<g font-family="Arial">{labels}</g>
<rect x="{width-275}" y="70" width="220" height="70" fill="#fff" stroke="#d8e0e8"/><line x1="{width-255}" y1="94" x2="{width-220}" y2="94" stroke="#0b5cab" stroke-width="3"/><text x="{width-210}" y="98" font-family="Arial" font-size="12" fill="#334">Aktual</text><line x1="{width-255}" y1="120" x2="{width-220}" y2="120" stroke="#0f766e" stroke-width="3"/><text x="{width-210}" y="124" font-family="Arial" font-size="12" fill="#334">Proyeksi OLS</text></svg>'''
    (OUT_DIR / "grafik_proyeksi_total.svg").write_text(svg, encoding="utf-8")


def write_workbook(result: dict, fitted: pd.DataFrame, projection: pd.DataFrame, gap: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook(); sheets = []
    for name, df in [("Input Aktual", fitted), ("Statistik OLS", pd.DataFrame([{"metrik":k,"nilai":v} for k,v in result.items()])), ("Dashboard Resmi", OFFICIAL_DASHBOARD), ("Proyeksi Tahunan", projection), ("Gap 2031", gap)]:
        ws = wb.active if not sheets else wb.create_sheet(name); ws.title = name; sheets.append(ws)
        for row in dataframe_to_rows(df, index=False, header=True): ws.append(row)
    notes = wb.create_sheet("Catatan")
    for n in ["Titik proyeksi tahunan disetarakan pada bulan Juli setiap tahun.", "Regresi OLS memakai total nasional: Juni 2025, Mei 2026, Juni 2026, Juli 2026.", "Komposisi per jenjang memakai proporsi angka tersedia dashboard resmi Juli 2026.", "Total time-series Juli 2026 adalah 27.837, sedangkan total komposisi dashboard resmi adalah 27.820.", "Data BKN 2025 tidak memuat usia/tanggal lahir; proyeksi pensiun presisi tidak dihitung."]:
        notes.append([n])
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+2, 42)
    for ws in [wb["Input Aktual"], wb["Dashboard Resmi"], wb["Proyeksi Tahunan"], wb["Gap 2031"]]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)): cell.number_format = "#,##0"
    for cell in wb["Gap 2031"]["E"][1:]: cell.number_format = "0.0%"
    ws = wb["Proyeksi Tahunan"]; chart = LineChart(); chart.title = "Proyeksi Total Nasional"; chart.y_axis.title = "Orang"; chart.x_axis.title = "Tahun"
    chart.add_data(Reference(ws, min_col=3, max_col=5, min_row=1, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)); chart.height = 8; chart.width = 18; ws.add_chart(chart, "N2")
    wb.save(OUT_DIR / "proyeksi_kebutuhan_sdm_arsiparis.xlsx")


def main() -> None:
    result, fitted = fit_ols(OBSERVATIONS); projection = build_projection(result); gap = build_gap_2031(projection)
    fitted.to_csv(OUT_DIR/"input_aktual_dan_fitted.csv", index=False, encoding="utf-8-sig")
    projection.to_csv(OUT_DIR/"tabel_proyeksi_tahunan_2026_2031.csv", index=False, encoding="utf-8-sig")
    OFFICIAL_DASHBOARD.to_csv(OUT_DIR/"angka_dashboard_resmi_juli_2026.csv", index=False, encoding="utf-8-sig")
    gap.to_csv(OUT_DIR/"gap_2031_vs_kebutuhan_ideal.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame([{"metrik":k, "nilai":v} for k,v in result.items()]).to_csv(OUT_DIR/"statistik_regresi_ols.csv", index=False, encoding="utf-8-sig")
    write_markdown(result, projection, gap); write_svg_chart(projection); write_workbook(result, fitted, projection, gap)
    print(f"Output tersimpan di: {OUT_DIR}"); print(pd.DataFrame([{"metrik":k, "nilai":v} for k,v in result.items()]).to_string(index=False)); print(projection.to_string(index=False)); print(gap.to_string(index=False))

if __name__ == "__main__":
    main()

